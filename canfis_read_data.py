import openpyxl
import pandas as pd
from pathlib import Path
import numpy as np
import os

def loadData(folder, filename, inputNo, header = True, slNo = 0):
    n,e = os.path.splitext(filename)
    if e == '.xlsx':
        xlsx_file = Path(folder, filename)
        wb_obj = openpyxl.load_workbook(xlsx_file) 

        # Read the active sheet:
        sheet = wb_obj.active
        print("Number of Rows:" + str(sheet.max_row), "Number of Columns:" + str(sheet.max_column))

        # col_names = []
        # for column in sheet.iter_cols(1, sheet.max_column):
        #     col_names.append(column[0].value)
            
        # print(col_names)
        # traindata = {}
        traininput = []
        traintarget = []
        colNames = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0 and header:
                j = 0
                while j < sheet.max_column:
                    colNames.append(row[j])
                    j = j + 1
            else:
                traininput.append([])
                traintarget.append([])
                j = 0
                while j < sheet.max_column:
                    if j >= slNo and j < inputNo+slNo:
                        traininput[-1].append(row[j])
                    elif j >= inputNo+slNo:
                        traintarget[-1].append(row[j])
                    # traindata[col_names[j]].append(row[j])
                    j = j + 1
        # print("Train Inputs:" + str(traininput))
        # print("Train Targets:" + str(traintarget))
        return [traininput, traintarget]
    if e == '.csv':
        data = pd.read_csv(os.path.join(folder,filename))
        inputs = data.iloc[:,0:inputNo]
        outputs = data.iloc[:,inputNo:]
        return [inputs, outputs]